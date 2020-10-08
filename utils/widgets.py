import streamlit as st
import openpyxl
import pandas as pd

st.set_option('deprecation.showfileUploaderEncoding', False)

def selectbox_without_default(label, options, blocker):
    """
    Wrapper for st.selectbox with empty string as default.
    In the app the default appears as specified in the
    parameter `blocker`.
    """

    options = [''] + options
    format_func = lambda x: blocker if x == '' else x
    return st.selectbox(label, options, format_func=format_func)


def excel_file_uploader(label):
    """
    Wrapper widget for reading data from excel file
    """

    # Load workbook
    file_buffer = st.file_uploader(label, type='xlsx')
    if not file_buffer:
        st.stop()
    wb = openpyxl.load_workbook(file_buffer, data_only=True)

    if len(wb.sheetnames) > 1:
        # Let the user choose worksheet
        worksheet = selectbox_without_default("Choose worksheet", wb.sheetnames,
                                              "Choose a worksheet")
        if not worksheet:
            st.stop()
        ws = wb.get_sheet_by_name(worksheet)
    else:
        # Get the only present worksheet
        ws = wb.worksheets[0]

    # Specify header if available
    header = st.checkbox("Header?", value=True)
    values = ws.values
    if header:
        columns = next(values)[0:]
        data = pd.DataFrame(values, columns=columns)
    else:
        data = pd.DataFrame(values)

    return data

